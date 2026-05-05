"""
Radar GovTech Brasil 2025/2026 — DLG
Gera: Radar_GovTechs_Brasil_2026.xlsx  +  index.html
Uso:  python govtech_excel.py
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# DADOS — edite aqui para atualizar Excel e HTML juntos
# ══════════════════════════════════════════════════════════════

VERSION = "v7"
DATA_REF = "Maio 2026"

COMPANIES = [
    # ── GESTÃO MUNICIPAL ──
    {"seg":"gestao", "name":"Aprova Digital",    "subseg":"Gestão Pública Municipal",   "porte":"Startup / Scale-up",  "loc":"São Paulo - SP",        "fund":"2017",  "presenca":"120+ cidades",            "receita":"R$ 22,5M (Astella+BB)",  "nota":"Aporte Seed de R$22,5M (US$4M) liderado por Astella e VOX Capital (CVC do Banco do Brasil), com CAF e Endeavor. 120+ cidades. TOP Open Startups 2024. Maior aporte da história das GovTechs da AL.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"GovDigital",        "subseg":"Gestão Pública Municipal",   "porte":"Startup / Scale-up",  "loc":"Jaraguá do Sul - SC",   "fund":"~2020", "presenca":"5 estados | 600+ serv.",  "receita":"N/D (bootstrap)",  "nota":"App SaaS municipal customizável. 600+ serviços digitalizados. CEO: Elias Raasch. Expansão nacional sem VC. Referência SC.", "alerta":"Watch — expansão nacional",    "isNew":False},
    {"seg":"gestao", "name":"Desenvolve Cidade", "subseg":"Gestão Pública Municipal",   "porte":"Startup",             "loc":"Campinas/RJ",            "fund":"~2014", "presenca":"N/D",                     "receita":"N/D",              "nota":"TOP 2 Ranking 100 Open Startups 2024 (GovTech).", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"IP Inovação",       "subseg":"Gestão Pública Municipal",   "porte":"Startup",             "loc":"Belém - PA",             "fund":"~2018", "presenca":"N/D",                     "receita":"N/D",              "nota":"TOP 3 Open Startups 2024. Representante do ecossistema Norte do Brasil.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"LICI GovTech",      "subseg":"Smart Cities / Gestão",      "porte":"Startup",             "loc":"Belo Horizonte - MG",   "fund":"2018",  "presenca":"Municípios | 2M+ cid.",   "receita":"N/D",              "nota":"Plataforma CHESI: 160+ indicadores ODS + fontes de financiamento + governança. Parceria SEDE-MG Cidades do Futuro.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"Quasar",            "subseg":"Gestão Municipal / Fiscal",  "porte":"Startup",             "loc":"Belo Horizonte - MG",   "fund":"2019",  "presenca":"Municípios MG",           "receita":"N/D",              "nota":"Alvarás digitais: reduz 90% do prazo de emissão. 2º DemoDay BrazilLAB 2023. Selecionada PBH Inova (2 desafios).", "alerta":"Watch — fiscal municipal",     "isNew":False},
    {"seg":"gestao", "name":"GESUAS",            "subseg":"Assistência Social / SUAS",  "porte":"Startup / PME",       "loc":"Viçosa - MG",           "fund":"~2016", "presenca":"168 mun | 3,5M pessoas", "receita":"N/D",              "nota":"Prontuário eletrônico SUAS. 1,2M famílias | 6,5k trabalhadores. Líder IE GovTech 2020. Vertical exclusivo sem concorrente direto.", "alerta":"Vertical exclusivo mapeado",   "isNew":False},
    {"seg":"gestao", "name":"Prosas",            "subseg":"Participação Cidadã",        "porte":"Startup",             "loc":"São Paulo - SP",        "fund":"~2016", "presenca":"220k+ usuários",          "receita":"R$4M (KPTL)",      "nota":"Portfólio Fundo GovTech KPTL. Gestão digital de editais socioculturais e transparência pública.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"Lemobs",            "subseg":"Smart Cities / Mobilidade",  "porte":"Startup",             "loc":"Rio de Janeiro - RJ",   "fund":"~2015", "presenca":"N/D",                     "receita":"N/D",              "nota":"Origem COPPE/UFRJ. Selo GovTech BrazilLAB. Mobilidade urbana e cidades inteligentes.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"Sipremo",           "subseg":"Infraestrutura / Clima",     "porte":"Startup",             "loc":"Ribeirão Preto - SP",   "fund":"~2018", "presenca":"Defesa Civil + seguros",  "receita":"R$200k (2022)",    "nota":"IA para previsão antecipada de desastres naturais. Dados NASA + bases nacionais. Alerta SMS via Defesa Civil. COP26 selecionada.", "alerta":"Watch — clima/risco",          "isNew":False},
    {"seg":"gestao", "name":"i4Sea",             "subseg":"Infraestrutura / Clima",     "porte":"Startup",             "loc":"Santos - SP",           "fund":"~2018", "presenca":"N/D",                     "receita":"R$ 7,5M aporte",   "nota":"Previsões microclimáticas para portos e setor elétrico. Captou R$7,5M do Fundo GovTech. Resiliência climática.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"GRTS Digital",      "subseg":"Relações Trabalhistas",      "porte":"Startup",             "loc":"São Paulo - SP",        "fund":"2019",  "presenca":"N/D",                     "receita":"R$ 3,4M aporte",   "nota":"Digitalização da gestão de relações sindicais. Captou R$3,4M do Fundo GovTech KPTL.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"IPQ Tecnologia",    "subseg":"Segurança Pública",          "porte":"PME / Scale-up",      "loc":"Curitiba - PR",         "fund":"~2010", "presenca":"250+ municípios",         "receita":"N/D",              "nota":"Monitoramento e recuperação de veículos. 7.000+ veículos recuperados em 2024. Redução de 15,5% em furtos.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"Atech",             "subseg":"Segurança Pública / Defesa", "porte":"Médio Porte",         "loc":"São Paulo - SP",        "fund":"1999",  "presenca":"Nacional / Defesa",       "receita":"N/D",              "nota":"Empresa Estratégica de Defesa (MDefesa). Plataformas de Comando & Controle e gestão de crises.", "alerta":"",                             "isNew":False},
    {"seg":"gestao", "name":"Governar",          "subseg":"ERP Municipal / SIAFIC",     "porte":"Startup / Scale-up",  "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"Municípios (SIAFIC)",     "receita":"N/D",              "nota":"ERP municipal integrado 100% cloud: SIAFIC + saúde + educação + assistência social. Conformidade SIAFIC obrigatória impulsiona substituição de legados.", "alerta":"Watch — SIAFIC driver",        "isNew":True},
    {"seg":"gestao", "name":"Qiatech",           "subseg":"Convênios / Transferências", "porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"Municípios (convênios)",  "receita":"N/D",              "nota":"Gestão de convênios com IA: integra TransfereGov, Simec, SISMOB, FNS. Vertical exclusivo de captação e gestão de transferências federais para municípios.", "alerta":"Watch — convênios nicho virgem","isNew":True},
    {"seg":"gestao", "name":"MuniScore",         "subseg":"Inteligência Fiscal",        "porte":"Startup",             "loc":"Belo Horizonte - MG",   "fund":"2026",  "presenca":"5.570 municípios (dados)","receita":"N/D",              "nota":"Score fiscal para todos os 5.570 municípios brasileiros. Fundada 2026. Plataforma B2B/B2G de avaliação de saúde fiscal e creditícia municipal.", "alerta":"Watch — dado fiscal municipal", "isNew":True},
    {"seg":"gestao", "name":"Colab",             "subseg":"Participação Cidadã / Gestão Pública","porte":"Scale-up",         "loc":"Brasil",                "fund":"2013",  "presenca":"20M+ pessoas impactadas", "receita":"R$3-7M (Fundo GovTech)", "nota":"Plataforma de gestão pública e participação cidadã: zeladoria urbana, CRM prefeitura-cidadão e marketplace de serviços públicos digitais. 1º investimento da KPTL em 2014. Follow-on pelo Fundo GovTech (KPTL+Cedro) em 2023. Co-invest EDP.", "alerta":"Portfólio KPTL+Cedro",         "isNew":True},
    {"seg":"gestao", "name":"Fractal",           "subseg":"Gestão Hídrica / Infraestrutura","porte":"Startup / Scale-up",  "loc":"Brasil",                "fund":"2010",  "presenca":"Setor elétrico + agro",   "receita":"N/D (KPTL 2019)",  "nota":"SaaS hidrológico integrado: previsão de fluxo de água e gestão de bacias hidrográficas. 80 anos de dados históricos. Clientes: CPFL, Statkraft. Aplicações em energia, agronegócio e seguros. Investimento KPTL 2019 (pré-Fundo GovTech).", "alerta":"Portfólio KPTL",              "isNew":True},

    # ── SAÚDE PÚBLICA ──
    {"seg":"saude",  "name":"OM30",              "subseg":"Atenção Primária / Suite",   "porte":"PME Estabelecida",    "loc":"São Paulo - SP",        "fund":"~2005", "presenca":"Municípios SP",           "receita":"N/D",              "nota":"Suite municipal: Saúde Simples + Educação + DadosGov + Mobilidade + Doc+. 350 colaboradores. Meta: 2x municípios SP até 2030. IA reduz 40% tempo de consultas SUS.", "alerta":"Target roll-up",               "isNew":False},
    {"seg":"saude",  "name":"Olostech",          "subseg":"Atenção Primária / APS",     "porte":"PME Estabelecida",    "loc":"Joinville - SC",        "fund":"1992",  "presenca":"100+ municípios",         "receita":"N/D",              "nota":"Pioneira SaaS cloud em APS (1998). 91% dos clientes melhoraram indicadores Previne Brasil 2023.", "alerta":"Target roll-up",               "isNew":False},
    {"seg":"saude",  "name":"Vivver Sistemas",   "subseg":"APS + Educação Municipal",   "porte":"PME Estabelecida",    "loc":"Belo Horizonte - MG",   "fund":"1999",  "presenca":"Municípios MG + NE",      "receita":"N/D",              "nota":"26 anos de operação. Selos Ouro/Prata Bora Vacinar (mar/2025). Filiais SP e PB. Foco saúde pública + educação municipal integradas.", "alerta":"Target roll-up",               "isNew":False},
    {"seg":"saude",  "name":"Pública Tecnologia","subseg":"Prontuário Eletrônico APS",  "porte":"PME",                 "loc":"N/D",                   "fund":"N/D",   "presenca":"Municípios",              "receita":"N/D",              "nota":"Prontuário eletrônico certificado SBIS/CFM. Assinatura ICP-Brasil. Integração e-SUS.", "alerta":"",                             "isNew":False},
    {"seg":"saude",  "name":"ImpulsoGov",        "subseg":"Analytics / BI para APS",    "porte":"Startup early-stage", "loc":"São Paulo - SP",        "fund":"~2020", "presenca":"Municípios via SUS",      "receita":"US$ 100k",         "nota":"BI para gaps de atenção básica preventiva. Cofundada por ex-Harvard Kennedy School. Reconhecida pela CGU. Impulso Previne.", "alerta":"",                             "isNew":False},
    {"seg":"saude",  "name":"UpCities",          "subseg":"Integração e-SUS / Telecare","porte":"Startup",             "loc":"São Paulo - SP",        "fund":"~2019", "presenca":"Municípios (expansão)",   "receita":"N/D",              "nota":"Integradora e-SUS/RNDS. Prontuário unificado, telecare. App offline para agentes comunitários de saúde.", "alerta":"",                             "isNew":False},
    {"seg":"saude",  "name":"ToGov",             "subseg":"Analytics APS / Educação",   "porte":"Startup",             "loc":"Santa Luzia - MG",      "fund":"2020",  "presenca":"Municípios MG e outros",  "receita":"N/D",              "nota":"Analytics Previne Brasil: gaps, captação de recursos, indicadores APS. 3ª DemoDay BrazilLAB 2023.", "alerta":"",                             "isNew":False},
    {"seg":"saude",  "name":"MedBolso",          "subseg":"Escalas / Anti-fraude Saúde","porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"~2020", "presenca":"UBS / Municípios",        "receita":"N/D",              "nota":"Gestão de escalas e plantões para saúde pública municipal. Anti-fraude em folha de pagamento. Top 30 Sebrae 2024. Acelerada GovTech Pará.", "alerta":"Watch — anti-fraude saúde",    "isNew":True},

    # ── EDUCAÇÃO PÚBLICA ──
    {"seg":"educ",   "name":"Portabilis",        "subseg":"Gestão Escolar (SIE)",       "porte":"Scale-up / PME",      "loc":"Içara - SC",            "fund":"2009",  "presenca":"140+ mun | 700k+ alunos","receita":"Aporte Yunus R$2M", "nota":"Mantenedora i-Educar (software público federal). Parceria MEC e Fundação Lemann. Meta: 1,5M alunos até 2027.", "alerta":"Target / impacto",             "isNew":False},
    {"seg":"educ",   "name":"Vivver Sistemas",   "subseg":"Gestão Escolar (SIE)",       "porte":"PME Estabelecida",    "loc":"Belo Horizonte - MG",   "fund":"1999",  "presenca":"Municípios MG + NE",      "receita":"N/D",              "nota":"Cliente Sorocaba (SP): avançou C+ → B no IEGM 2025 (Saúde e Educação). 26 anos no segmento.", "alerta":"",                             "isNew":False},
    {"seg":"educ",   "name":"Educ21",            "subseg":"Gestão Escolar com IA",      "porte":"PME / Startup",       "loc":"N/D",                   "fund":"N/D",   "presenca":"Municípios",              "receita":"N/D",              "nota":"Gestão acadêmica, administrativa e financeira com IA integrada. Dados institucionais limitados — oportunidade de due diligence.", "alerta":"",                             "isNew":False},
    {"seg":"educ",   "name":"Geekie",            "subseg":"Aprendizagem IA / Híbrida",  "porte":"Scale-up",            "loc":"São Paulo - SP",        "fund":"2011",  "presenca":"5k esc | 12M alunos",    "receita":"N/D",              "nota":"415k alunos rede estadual SP (gratuito). Parceria MEC desde 2014. IA + ensino híbrido. Potencial alvo M&A.", "alerta":"",                             "isNew":False},
    {"seg":"educ",   "name":"Educacross",        "subseg":"Gamificação + IA Educacional","porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"Rede estadual GO",        "receita":"N/D",              "nota":"Selecionada Governo de Goiás. Gamificação + IA em Português e Matemática para redes públicas.", "alerta":"",                             "isNew":False},
    {"seg":"educ",   "name":"ToGov",             "subseg":"Analytics Educacionais",     "porte":"Startup",             "loc":"Santa Luzia - MG",      "fund":"2020",  "presenca":"Municípios MG e outros",  "receita":"N/D",              "nota":"Monitoramento IDEB, IEGM, captação de recursos. Camada analítica sobre SIEs existentes. BrazilLAB 2023.", "alerta":"",                             "isNew":False},
    {"seg":"educ",   "name":"i-Educar",          "subseg":"Open Source / Software Público","porte":"Software Público",  "loc":"Itajaí - SC (origem)", "fund":"2008",  "presenca":"80+ prefeituras ativas",  "receita":"Gratuito",         "nota":"Software público federal mantido por Portabilis. Monte Alegre (RN) economizou R$2,4M. Parceria MEC + MP.", "alerta":"Referência pública",           "isNew":False},
    {"seg":"educ",   "name":"Jovens Gênios",     "subseg":"EdTech Gamificada / IA Educacional","porte":"Startup",          "loc":"Rio de Janeiro - RJ",   "fund":"N/D",   "presenca":"50+ municípios | est. SP","receita":"R$11,8M (Fundo GovTech)", "nota":"Plataforma adaptativa de aprendizagem e avaliação gamificada. 90% dos alunos em escolas públicas. 83% dos usuários em rede pública. Meta: 10M alunos até 2030. Rodada seed R$11,8M liderada pelo Fundo GovTech (KPTL+Cedro) com DOMO.VC, Criabiz Ventures e Rosey Ventures (Grupo Marista). Mar/2026.", "alerta":"Portfólio KPTL+Cedro",         "isNew":True},

    # ── IA PARA GOVERNO ──
    {"seg":"ia",     "name":"GovTools",          "subseg":"Agentes RPA / WhatsApp",     "porte":"Startup",             "loc":"Porto Alegre - RS",     "fund":"2024",  "presenca":"200+ municípios",         "receita":"Aporte Ventiur+DOMO","nota":"Agentes IA via WhatsApp. Opera sistemas municipais sem integração tradicional. 200+ municípios em <12 meses. Modelo viral, alta tração.", "alerta":"M&A Watch",                   "isNew":False},
    {"seg":"ia",     "name":"Nuveo",             "subseg":"Document AI / Ultra OCR",    "porte":"Startup",             "loc":"SP / Campina Grande - PB","fund":"2016", "presenca":"Setor público e privado", "receita":"N/D",              "nota":"Ultra OCR®: automação de documentos públicos. Contratos, invoices, documentos de identidade. Gov + corporativo.", "alerta":"",                             "isNew":False},
    {"seg":"ia",     "name":"CPQD",              "subseg":"IA P&D / Gov Digital",       "porte":"Instituto P&D",       "loc":"Campinas - SP",         "fund":"N/D",   "presenca":"Governo Federal",         "receita":"R$ 390M contrato", "nota":"Projeto INSPIRE com MGI (R$390M, 4 anos). IA personalizada para serviços de governo digital: chatbots, recomendação, acessibilidade.", "alerta":"Parceria estratégica MGI",     "isNew":False},
    {"seg":"ia",     "name":"Horus Smart Det.",  "subseg":"Fiscalização / Drones / IA", "porte":"Scale-up",            "loc":"Florianópolis - SC",    "fund":"~2016", "presenca":"100+ gov + privado",      "receita":"N/D",              "nota":"Drones + satélites + visão computacional: obras irregulares, desmatamento, infraestrutura. Cliente: Florianópolis. 9 anos de operação.", "alerta":"",                             "isNew":False},
    {"seg":"ia",     "name":"Geopixel",          "subseg":"GeoInteligência Fiscal",     "porte":"PME",                 "loc":"Curitiba - PR",         "fund":"~2010", "presenca":"100+ municípios",         "receita":"N/D",              "nota":"Geoprocessamento + IA para IPTU, ITBI, ISS. Reforma Tributária (LC 214/2025) cria janela de demanda direta e imediata.", "alerta":"Oportunidade fiscal 2025",     "isNew":False},
    {"seg":"ia",     "name":"Intelicity",        "subseg":"IA Infraestrutura Urbana",   "porte":"Startup / Scale-up",  "loc":"São Paulo - SP",        "fund":"~2019", "presenca":"Sabesp + municípios",     "receita":"N/D",              "nota":"Câmeras em veículos + visão computacional: mapeia pavimento, rachaduras. Sabesp: detecção de vazamentos de água.", "alerta":"",                             "isNew":False},
    {"seg":"ia",     "name":"MinutaIA (jAI)",    "subseg":"IA Jurídica / Judicial",     "porte":"Startup",             "loc":"Belo Horizonte - MG",   "fund":"2025",  "presenca":"7 tribunais | 100k+ usr", "receita":"N/D",              "nota":"10M+ minutas geradas. TJRS, STM, TRE-PB/BA, Tocantins. CNJ Resolução 615/2025. Exportou para a Argentina.", "alerta":"Watch — tração judicial",      "isNew":False},
    {"seg":"ia",     "name":"Tolky",             "subseg":"IA Atendimento ao Cidadão",  "porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"~2023", "presenca":"Contratos gov / CNJ",     "receita":"N/D",              "nota":"Multichannel (WhatsApp, chat, email) + processos end-to-end. ABES CSC 2025. Contrato CPSI CNJ.", "alerta":"Selecionada CPSI CNJ",         "isNew":False},
    {"seg":"ia",     "name":"Gove",              "subseg":"Analytics Financeiro Municipal","porte":"Startup / Scale-up","loc":"São Paulo - SP",        "fund":"2018",  "presenca":"SP, MG, RS, SC, ES+",    "receita":"R$ 8M (Astella)",  "nota":"Analytics IA para finanças públicas: identifica ineficiências em receitas/despesas municipais. Aporte Astella. BrazilLAB Turma 7.", "alerta":"",                             "isNew":False},
    {"seg":"ia",     "name":"Hub Esfera",        "subseg":"IA Analytics / Gestão",      "porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"N/D",                     "receita":"N/D",              "nota":"Plataforma IA para gestão municipal: analytics preditivo + preventivo + prescritivo. Integra dados orçamentários, operacionais e sociais para apoio à decisão do gestor público.", "alerta":"Watch",                        "isNew":True},
    {"seg":"ia",     "name":"Kinebot",           "subseg":"IA Ergonomia / Saúde do Trabalhador","porte":"Startup",          "loc":"Curitiba - PR",         "fund":"N/D",   "presenca":"Indústria + gov",          "receita":"R$3M (Fundo GovTech)","nota":"IA para análises ergonômicas e psicossociais via visão computacional. Clientes: Marfrig, P&G, Electrolux. 8º investimento do Fundo GovTech (KPTL+Cedro), R$3M em fev/2026. Foco em saúde do trabalhador para setor público e privado. Expansão global planejada.", "alerta":"Portfólio KPTL+Cedro",         "isNew":True},

    # ── PROCURADORIAS ──
    {"seg":"proc",   "name":"Attus Procuradoria Digital","subseg":"IA Contencioso + Consultivo (PGE/PGM)","porte":"Scale-up","loc":"São Ludgero, SC (Eloware)","fund":"~2018","presenca":"PGE-SP, PGE-BA, PGE-PA, PGM-Palmas","receita":"SaaS Oracle Cloud","nota":"4,3M+ processos analisados. 73% classificados automaticamente (98,5% acurácia). 642k+ peças geradas por IA. Referência nacional pós PGE-SP.", "alerta":"M&A Watch — líder do segmento","isNew":False},
    {"seg":"proc",   "name":"Eicon / Giex",      "subseg":"Dívida Ativa / Execução Fiscal","porte":"PME",              "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"PGE e PGM — múltiplos estados","receita":"N/D",        "nota":"Giex: perfil completo do devedor. Integração direta com TJs. Automação de CDA, penhoras e prazos. iCad: alvarás e licenças digitais (concorrente da Quasar). +15 produtos para governo.", "alerta":"",                             "isNew":False},
    {"seg":"proc",   "name":"PGMNET",            "subseg":"Gestão Jurídica Municipal + RPA","porte":"PME",             "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"Procuradorias municipais","receita":"N/D",              "nota":"RPA integrado para peticionamento automático em qualquer TJ do país. Interface co-desenvolvida com procuradores municipais.", "alerta":"",                             "isNew":False},
    {"seg":"proc",   "name":"Aetos Tech",        "subseg":"IA Anti-fraude Contratos (Prisma)","porte":"Startup",       "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"MPRJ (CPSI 2024)",        "receita":"Contrato CPSI",    "nota":"Prisma: detecção de irregularidades em contratos públicos e lavagem de dinheiro. Cruzamento de bases abertas por IA. Selecionada CPSI MPRJ.", "alerta":"Contrato CPSI",                "isNew":False},
    {"seg":"proc",   "name":"Sumé Tecnologia",   "subseg":"Monitoramento Políticas Públicas (Sonar)","porte":"Startup", "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"MPRJ (CPSI 2024)",        "receita":"Contrato CPSI",    "nota":"Sonar: controle externo automatizado do executivo municipal. Monitora 92 municípios do RJ via MPRJ. Selecionada CPSI MPRJ.", "alerta":"Contrato CPSI",                "isNew":False},
    {"seg":"proc",   "name":"Prodata Gestão Estratégica","subseg":"Módulo Procuradoria / ERP Municipal","porte":"PME",  "loc":"Goiânia - GO",          "fund":"N/D",   "presenca":"Municípios Centro-Oeste", "receita":"N/D",              "nota":"Módulo procuradoria integrado ao ERP municipal Prodata. Gestão de contencioso, consultivo e execução fiscal para prefeituras de pequeno e médio porte.", "alerta":"",                             "isNew":False},

    # ── LICITAÇÕES ──
    {"seg":"lic",    "name":"Portal de Compras Públicas","subseg":"Marketplace Licitações Eletrônicas","porte":"Scale-up / PME","loc":"Belo Horizonte - MG","fund":"2016","presenca":"4.300+ entes | 600k+ forn.","receita":"R$450B+ TPV",  "nota":"40%+ dos municípios brasileiros. Fundadores: Leonardo + Bruno Ladeira (Ecustomize). 35% crescimento anual. Marketplace B2G de referência nacional.", "alerta":"Referência nacional",          "isNew":False},
    {"seg":"lic",    "name":"Licitar.digital",   "subseg":"SaaS Credenciamento Eletrônico","porte":"Startup",           "loc":"N/D (Brasil)",          "fund":"Jul/2019","presenca":"90k+ fornecedores",      "receita":"N/D",              "nota":"Única plataforma de credenciamento eletrônico automatizado do Brasil. CEO: Aniele H. Figueiredo. Diferencial regulatório: elimina burocracia de habilitação presencial.", "alerta":"",                             "isNew":False},
    {"seg":"lic",    "name":"StartGi",           "subseg":"CRM B2G para Fornecedores",  "porte":"Startup",             "loc":"São Paulo - SP",        "fund":"2015",  "presenca":"N/D",                     "receita":"N/D",              "nota":"Portfólio Fundo GovTech KPTL. CRM especializado na jornada de vendas B2G. Gestão de oportunidades, propostas e relacionamento com entes públicos.", "alerta":"",                             "isNew":False},
    {"seg":"lic",    "name":"BLL Compras",       "subseg":"Plataforma Licitações Eletrônicas","porte":"PME Estabelecida","loc":"São Paulo - SP",        "fund":"2008",  "presenca":"3.000+ entidades públicas","receita":"N/D",              "nota":"Bolsa de Licitações e Leilões — uma das maiores plataformas de licitações do Brasil. Mais consolidada, pré-startups. Referência histórica do mercado.", "alerta":"",                             "isNew":False},
    {"seg":"lic",    "name":"GoBuyer",           "subseg":"IA Contratos Pós-adjudicação","porte":"Startup",             "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"N/D",                     "receita":"N/D",              "nota":"Gestão do ciclo de vida de contratos públicos com IA: análise de cláusulas, alertas de prazo, auditoria automatizada. Segmento pós-licitação ainda sem líder claro.", "alerta":"Watch — nicho virgem",         "isNew":False},
    {"seg":"lic",    "name":"StartGov",          "subseg":"Licitações Buy-side para Órgãos Públicos","porte":"PME",     "loc":"N/D (Brasil)",          "fund":"N/D",   "presenca":"290+ órgãos públicos",    "receita":"N/D",              "nota":"Plataforma buy-side: gestão de licitações e contratos para entes públicos. 290+ clientes. Lei 14.133/2021 compliance. Distinto do StartGi (B2G para fornecedores privados).", "alerta":"Watch — buy-side gov",         "isNew":True},
]

APORTES = [
    {"tx":"Gove — Astella",                     "tipo":"Aporte Series A",   "valor":"R$ 8M",     "ano":2020, "data":"2020",      "seg":"IA Analytics Financeiro", "desc":"Maior aporte em govtech na época. Plataforma analytics financeiro para prefeituras. Expansão geográfica em curso."},
    {"tx":"Portabilis — Yunus Negócios Sociais", "tipo":"Aporte de impacto", "valor":"Até R$ 2M", "ano":2023, "data":"2023",      "seg":"Educação Pública",        "desc":"R$800k já desembolsados. Valida modelo de negócio social. Meta: 1,5M alunos até 2027."},
    {"tx":"ImpulsoGov — Mulago Foundation",      "tipo":"Aporte seed",       "valor":"US$ 100k",  "ano":2023, "data":"2023",      "seg":"Saúde — Analytics / APS", "desc":"Capital de impacto para BI de atenção básica preventiva. Cofundada por ex-Harvard Kennedy School."},
    {"tx":"GovTools — Ventiur + DOMO.VC",        "tipo":"Aporte early-stage","valor":"N/D",       "ano":2024, "data":"2024",      "seg":"IA Operacional Gov",      "desc":"200+ municípios em menos de 12 meses. Modelo viral via WhatsApp. Potencial alvo de aquisição."},
    {"tx":"i4Sea — Fundo GovTech KPTL",          "tipo":"Aporte",            "valor":"R$ 7,5M",   "ano":2024, "data":"2024",      "seg":"Infraestrutura / Clima",  "desc":"Previsões microclimáticas para portos e setor elétrico. Portfólio Fundo GovTech KPTL."},
    {"tx":"Augen — Biosolvit (exit)",            "tipo":"Exit estratégico",  "valor":"R$ 36–48M",  "ano":2024, "data":"out/2024",  "seg":"Infraestrutura / Água",    "desc":"1º exit do Fundo GovTech KPTL. Automação de tratamento de água. Earn-out 3 anos. Comprador: Biosolvit (Laércio Cosentino). Badesul apoiou."},
    {"tx":"Aprova Digital — Astella + BB/VOX",   "tipo":"Aporte Seed",       "valor":"R$ 22,5M",   "ano":2022, "data":"2022",      "seg":"Gestão Pública Municipal", "desc":"Rodada liderada por Astella e VOX Capital (CVC do Banco do Brasil), com CAF e Endeavor. Maior aporte da história das GovTechs da América Latina na época. 120+ cidades, 21M de brasileiros impactados."},
    {"tx":"Kinebot — Fundo GovTech (KPTL+Cedro)","tipo":"Aporte",            "valor":"R$ 3M",      "ano":2026, "data":"fev/2026",  "seg":"IA Ergonomia / Saúde Gov", "desc":"8º investimento do Fundo GovTech. IA para análises ergonômicas e psicossociais. Clientes: Marfrig, P&G, Electrolux. Expansão global planejada para 2026."},
    {"tx":"Jovens Gênios — Fundo GovTech+DOMO",  "tipo":"Aporte Seed",       "valor":"R$ 11,8M",   "ano":2026, "data":"mar/2026",  "seg":"Educação Pública",         "desc":"Rodada seed liderada pelo Fundo GovTech (KPTL+Cedro), com DOMO.VC, Criabiz Ventures e Rosey Ventures (Grupo Marista). 90% dos alunos em escolas públicas. Meta: 10M alunos até 2030."},
]

# ══════════════════════════════════════════════════════════════
# CONFIGURAÇÕES DE SEGMENTOS
# ══════════════════════════════════════════════════════════════

SEGMENTS = {
    "gestao": {"label": "Gestão Pública Municipal",      "emoji": "🏛️",  "color": "1F4E79"},
    "saude":  {"label": "Saúde Pública",                 "emoji": "🏥",  "color": "0D4F4A"},
    "educ":   {"label": "Educação Pública",              "emoji": "📚",  "color": "2E75B6"},
    "ia":     {"label": "IA para Governo",               "emoji": "🤖",  "color": "6A1B9A"},
    "proc":   {"label": "Procuradorias (PGM/PGE/MP)",   "emoji": "⚖️",  "color": "BF360C"},
    "lic":    {"label": "Licitações & Compras Públicas", "emoji": "📋",  "color": "2E7D32"},
}

# ══════════════════════════════════════════════════════════════
# GERADOR DE EXCEL
# ══════════════════════════════════════════════════════════════

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_excel(output_path):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Aba Resumo ──
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    ws.merge_cells("A1:B1")
    ws["A1"] = f"🇧🇷 Radar GovTech Brasil 2025/2026 — DLG · {DATA_REF} · {VERSION}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hex_fill("1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A3"] = "Segmento"
    ws["B3"] = "Empresas"
    for cell in [ws["A3"], ws["B3"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hex_fill("2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    row = 4
    total = 0
    unique = {c["name"] for c in COMPANIES}
    for seg, cfg in SEGMENTS.items():
        count = len([c for c in COMPANIES if c["seg"] == seg])
        total += count
        ws.cell(row, 1, f'{cfg["emoji"]} {cfg["label"]}').border = thin_border()
        ws.cell(row, 2, count).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).border = thin_border()
        row += 1

    ws.cell(row, 1, "TOTAL (incluindo aparições em múltiplos segs)").font = Font(bold=True)
    ws.cell(row, 1).fill = hex_fill("D6E4F0")
    ws.cell(row, 2, total).font = Font(bold=True)
    ws.cell(row, 2).fill = hex_fill("D6E4F0")
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    row += 1
    ws.cell(row, 1, f"Empresas únicas mapeadas").font = Font(bold=True)
    ws.cell(row, 2, len(unique)).font = Font(bold=True)
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    row += 2
    ws.cell(row, 1, f"Aportes & Investimentos registrados")
    ws.cell(row, 2, len(APORTES)).alignment = Alignment(horizontal="center")

    # ── Aba por segmento ──
    COLS = ["Nome", "Subsegmento", "Porte", "Localização", "Fundação", "Presença", "Capital / Receita", "Análise / Nota", "Alerta M&A", "Novo"]
    COL_W = [28, 32, 20, 24, 10, 28, 20, 60, 28, 6]

    for seg, cfg in SEGMENTS.items():
        safe_label = cfg["label"][:25].replace("/", "-")
        ws = wb.create_sheet(f'{cfg["emoji"]} {safe_label}')
        ws.sheet_view.showGridLines = False

        # cabeçalho
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        ws["A1"] = f'{cfg["emoji"]} {cfg["label"]} — Radar GovTech Brasil 2025/2026'
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = hex_fill(cfg["color"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for ci, (col, w) in enumerate(zip(COLS, COL_W), 1):
            c = ws.cell(2, ci, col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hex_fill("2E75B6")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        filtered = [c for c in COMPANIES if c["seg"] == seg]
        for ri, comp in enumerate(filtered, 3):
            row_data = [
                comp["name"], comp["subseg"], comp["porte"], comp["loc"],
                comp["fund"], comp["presenca"], comp["receita"], comp["nota"],
                comp["alerta"], "✓" if comp["isNew"] else ""
            ]
            bg = "FFFFFF" if ri % 2 == 1 else "F4F6F9"
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(ri, ci, val)
                c.fill = hex_fill(bg)
                c.border = thin_border()
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if ci == 1:
                    c.font = Font(bold=True)
                if ci == 9 and val:
                    c.font = Font(color="8B0000", bold=True)
                if ci == 10 and val:
                    c.font = Font(color="1A7A72", bold=True)
            ws.row_dimensions[ri].height = 48

    # ── Aba Aportes ──
    ws = wb.create_sheet("💰 Aportes & Investimentos")
    ws.sheet_view.showGridLines = False
    a_cols = ["Empresa / Investidor", "Tipo", "Valor", "Data", "Segmento", "Descrição"]
    a_widths = [32, 20, 14, 12, 24, 60]

    ws.merge_cells(f"A1:{get_column_letter(len(a_cols))}1")
    ws["A1"] = "💰 Aportes & Investimentos — Radar GovTech Brasil 2025/2026"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = hex_fill("E65100")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (col, w) in enumerate(zip(a_cols, a_widths), 1):
        c = ws.cell(2, ci, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hex_fill("E65100")
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, ap in enumerate(sorted(APORTES, key=lambda x: x["ano"]), 3):
        row_data = [ap["tx"], ap["tipo"], ap["valor"], ap["data"], ap["seg"], ap["desc"]]
        bg = "FFFFFF" if ri % 2 == 1 else "FFF3E0"
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(ri, ci, val)
            c.fill = hex_fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == 1:
                c.font = Font(bold=True)
            if "Exit" in ap["tipo"] and ci == 2:
                c.font = Font(color="8B0000", bold=True)
        ws.row_dimensions[ri].height = 48

    wb.save(output_path)
    print(f"  ✓ Excel salvo: {output_path}")

# ══════════════════════════════════════════════════════════════
# GERADOR DE HTML
# ══════════════════════════════════════════════════════════════

def make_html(output_path):
    companies_json = json.dumps(COMPANIES, ensure_ascii=False, indent=2)
    aportes_json   = json.dumps(sorted(APORTES, key=lambda x: x["ano"]), ensure_ascii=False, indent=2)

    unique_count = len({c["name"] for c in COMPANIES})

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Radar GovTech Brasil 2025/2026 — DLG</title>
<style>
  :root {{
    --blue-dark: #1F4E79;
    --blue-mid: #2E75B6;
    --blue-light: #D6E4F0;
    --teal-dark: #0D4F4A;
    --teal-mid: #1A7A72;
    --teal-light: #C8E6E4;
    --purple-dark: #4A148C;
    --purple-mid: #6A1B9A;
    --purple-light: #EDE7F6;
    --orange-dark: #7B3F00;
    --orange-mid: #E65100;
    --orange-fill: #FFF3E0;
    --amber-dark: #BF360C;
    --amber-fill: #FFF8E1;
    --lime-dark: #2E7D32;
    --lime-light: #F1F8E9;
    --slate-dark: #37474F;
    --slate-light: #ECEFF1;
    --red-fill: #FFEBEE;
    --red-dark: #8B0000;
    --gray-bg: #F4F6F9;
    --white: #FFFFFF;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: var(--gray-bg); color: #1a1a1a; font-size: 13px; }}

  .header {{ background: linear-gradient(135deg, var(--blue-dark) 0%, #2a5f8e 100%); color: white; padding: 28px 40px 22px; }}
  .header h1 {{ font-size: 22px; font-weight: 700; letter-spacing: 0.5px; margin-bottom: 4px; }}
  .header .sub {{ font-size: 12px; opacity: 0.75; letter-spacing: 1px; text-transform: uppercase; }}
  .header-meta {{ display: flex; gap: 32px; margin-top: 18px; flex-wrap: wrap; }}
  .kpi {{ background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2); border-radius: 8px; padding: 12px 20px; text-align: center; min-width: 110px; }}
  .kpi .num {{ font-size: 26px; font-weight: 800; line-height: 1; }}
  .kpi .lbl {{ font-size: 10px; opacity: 0.8; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}

  .controls {{ background: white; padding: 14px 40px; border-bottom: 1px solid #dde3ec; display: flex; gap: 10px; align-items: center; flex-wrap: wrap; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
  .search-box {{ flex: 1; min-width: 220px; max-width: 340px; border: 1.5px solid #c5d0e0; border-radius: 6px; padding: 7px 12px; font-size: 13px; outline: none; transition: border-color 0.2s; }}
  .search-box:focus {{ border-color: var(--blue-mid); }}
  .filter-btns {{ display: flex; gap: 6px; flex-wrap: wrap; align-items: center; }}
  .filter-btns .lbl {{ font-size: 11px; font-weight: 600; color: #666; text-transform: uppercase; letter-spacing: 0.5px; margin-right: 4px; }}
  .btn {{ padding: 5px 12px; border-radius: 20px; border: 1.5px solid transparent; font-size: 11px; font-weight: 600; cursor: pointer; transition: all 0.18s; white-space: nowrap; }}
  .btn:hover {{ opacity: 0.88; transform: translateY(-1px); }}
  .btn.active {{ box-shadow: 0 2px 8px rgba(0,0,0,0.2); }}
  .btn-all   {{ background: var(--blue-dark); color: white; border-color: var(--blue-dark); }}
  .btn-gestao {{ background: #E3EDF9; color: var(--blue-dark); border-color: #b8cde0; }}
  .btn-gestao.active {{ background: var(--blue-dark); color: white; }}
  .btn-saude  {{ background: var(--teal-light); color: var(--teal-dark); border-color: #9dcfca; }}
  .btn-saude.active  {{ background: var(--teal-dark); color: white; }}
  .btn-educ   {{ background: var(--blue-light); color: var(--blue-dark); border-color: #9dc4e0; }}
  .btn-educ.active   {{ background: var(--blue-mid); color: white; }}
  .btn-ia     {{ background: var(--purple-light); color: var(--purple-dark); border-color: #c5a8e0; }}
  .btn-ia.active     {{ background: var(--purple-mid); color: white; }}
  .btn-proc   {{ background: var(--amber-fill); color: var(--amber-dark); border-color: #e0c090; }}
  .btn-proc.active   {{ background: var(--amber-dark); color: white; }}
  .btn-lic    {{ background: var(--lime-light); color: var(--lime-dark); border-color: #aad4a0; }}
  .btn-lic.active    {{ background: var(--lime-dark); color: white; }}
  .btn-ma     {{ background: var(--orange-fill); color: var(--orange-dark); border-color: #e0b080; }}
  .btn-ma.active     {{ background: var(--orange-dark); color: white; }}
  .count-badge {{ background: rgba(255,255,255,0.3); border-radius: 10px; padding: 1px 6px; font-size: 10px; margin-left: 4px; }}

  .main {{ padding: 24px 40px 40px; }}
  .hidden {{ display: none !important; }}

  .section-header {{ display: flex; align-items: center; gap: 12px; margin: 28px 0 14px; padding-bottom: 8px; border-bottom: 2px solid; }}
  .section-header h2 {{ font-size: 15px; font-weight: 700; }}
  .section-header .pill {{ font-size: 11px; font-weight: 600; padding: 3px 10px; border-radius: 12px; background: rgba(0,0,0,0.08); }}

  .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(340px, 1fr)); gap: 14px; }}

  .card {{ background: white; border-radius: 10px; padding: 16px 18px 14px; border: 1.5px solid #e0e8f0; transition: transform 0.15s, box-shadow 0.15s; position: relative; overflow: hidden; }}
  .card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,0.10); }}
  .card::before {{ content: ''; position: absolute; left: 0; top: 0; bottom: 0; width: 4px; border-radius: 4px 0 0 4px; }}
  .card.gestao::before {{ background: var(--blue-dark); }}
  .card.saude::before  {{ background: var(--teal-mid); }}
  .card.educ::before   {{ background: var(--blue-mid); }}
  .card.ia::before     {{ background: var(--purple-mid); }}
  .card.proc::before   {{ background: var(--amber-dark); }}
  .card.lic::before    {{ background: var(--lime-dark); }}

  .card-top {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 8px; gap: 8px; }}
  .card-name {{ font-size: 14px; font-weight: 700; line-height: 1.2; }}
  .card.gestao .card-name {{ color: var(--blue-dark); }}
  .card.saude  .card-name {{ color: var(--teal-dark); }}
  .card.educ   .card-name {{ color: var(--blue-mid); }}
  .card.ia     .card-name {{ color: var(--purple-mid); }}
  .card.proc   .card-name {{ color: var(--amber-dark); }}
  .card.lic    .card-name {{ color: var(--lime-dark); }}

  .card-porte {{ font-size: 10px; font-weight: 600; padding: 3px 8px; border-radius: 10px; white-space: nowrap; flex-shrink: 0; }}
  .porte-startup {{ background: #E8F5E9; color: #1B5E20; }}
  .porte-scale   {{ background: #E3F2FD; color: #0D47A1; }}
  .porte-pme     {{ background: #FFF9C4; color: #795548; }}
  .porte-medio   {{ background: #FCE4EC; color: #880E4F; }}
  .porte-grande  {{ background: #ECEFF1; color: #37474F; }}
  .porte-publico {{ background: #E8EAF6; color: #283593; }}

  .card-seg {{ font-size: 10.5px; color: #666; margin-bottom: 8px; font-style: italic; }}
  .card-meta {{ display: flex; gap: 12px; margin-bottom: 9px; flex-wrap: wrap; }}
  .meta-item {{ display: flex; align-items: center; gap: 4px; font-size: 11px; color: #555; }}
  .icon {{ font-size: 12px; }}
  .card-nota {{ font-size: 12px; color: #333; line-height: 1.5; }}

  .card-ref-tag   {{ display: inline-block; margin-top: 4px; font-size: 10px; font-weight: 700; padding: 2px 8px; border-radius: 10px; background: #E8F5E9; color: #1B5E20; border: 1px solid #a5d6a7; }}
  .card-watch-tag {{ display: inline-block; margin-top: 4px; font-size: 10px; font-weight: 700; padding: 2px 8px; border-radius: 10px; background: var(--amber-fill); color: var(--amber-dark); border: 1px solid #ffcc80; }}
  .card-new-tag   {{ display: inline-block; margin-top: 4px; font-size: 10px; font-weight: 700; padding: 2px 8px; border-radius: 10px; background: var(--purple-light); color: var(--purple-dark); border: 1px solid #ce93d8; margin-left: 4px; }}

  /* Aportes */
  .ma-table-wrap {{ overflow-x: auto; }}
  .ma-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  .ma-table th {{ background: var(--orange-dark); color: white; padding: 10px 14px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .ma-table td {{ padding: 10px 14px; border-bottom: 1px solid #f0e8e0; vertical-align: top; }}
  .ma-table tr:hover td {{ background: #fff8f0; }}
  .ma-tipo {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: 600; }}
  .tipo-aporte {{ background: #E3F2FD; color: #0D47A1; }}
  .tipo-exit   {{ background: #FCE4EC; color: #880E4F; }}

  .footer {{ text-align: center; padding: 24px; color: #999; font-size: 11px; border-top: 1px solid #e0e8f0; margin-top: 20px; }}

  /* Login */
  #login-overlay {{ position: fixed; inset: 0; background: linear-gradient(135deg, #1F4E79 0%, #2a5f8e 100%); display: flex; align-items: center; justify-content: center; z-index: 9999; }}
  #login-box {{ background: #fff; border-radius: 12px; padding: 48px 40px; width: 100%; max-width: 380px; box-shadow: 0 8px 32px rgba(0,0,0,0.25); text-align: center; }}
  #login-box h2 {{ color: #1F4E79; font-size: 20px; font-weight: 700; margin-bottom: 6px; }}
  #login-box p {{ color: #666; font-size: 13px; margin-bottom: 28px; }}
  #login-box input {{ width: 100%; padding: 12px 16px; border: 1.5px solid #ccc; border-radius: 8px; font-size: 15px; outline: none; margin-bottom: 12px; transition: border-color 0.2s; }}
  #login-box input:focus {{ border-color: #2E75B6; }}
  #login-box button {{ width: 100%; padding: 12px; background: #1F4E79; color: #fff; border: none; border-radius: 8px; font-size: 15px; font-weight: 600; cursor: pointer; transition: background 0.2s; }}
  #login-box button:hover {{ background: #2E75B6; }}
  #login-error {{ color: #c0392b; font-size: 13px; margin-top: 10px; display: none; }}
  #login-logo {{ font-size: 13px; color: #999; margin-top: 24px; }}
</style>
</head>
<body>

<div class="header">
  <h1>🇧🇷 Radar GovTech Brasil 2025/2026</h1>
  <div class="sub">DLG · Mapeamento M&amp;A — {DATA_REF} · {VERSION}</div>
  <div class="header-meta">
    <div class="kpi"><div class="num" id="kpi-total">{unique_count}</div><div class="lbl">Players mapeados</div></div>
    <div class="kpi"><div class="num">8</div><div class="lbl">Segmentos</div></div>
  </div>
</div>

<div class="controls">
  <input class="search-box" id="searchBox" type="text" placeholder="🔍  Buscar empresa, segmento, nota...">
  <div class="filter-btns">
    <span class="lbl">Filtro:</span>
    <button class="btn btn-all active" onclick="setFilter('all',this)">Todos <span class="count-badge" id="cnt-all">{unique_count}</span></button>
    <button class="btn btn-gestao" onclick="setFilter('gestao',this)">Gestão Municipal <span class="count-badge" id="cnt-gestao"></span></button>
    <button class="btn btn-saude"  onclick="setFilter('saude',this)">Saúde <span class="count-badge" id="cnt-saude"></span></button>
    <button class="btn btn-educ"   onclick="setFilter('educ',this)">Educação <span class="count-badge" id="cnt-educ"></span></button>
    <button class="btn btn-ia"     onclick="setFilter('ia',this)">IA Gov <span class="count-badge" id="cnt-ia"></span></button>
    <button class="btn btn-proc"   onclick="setFilter('proc',this)">Procuradorias <span class="count-badge" id="cnt-proc"></span></button>
    <button class="btn btn-lic"    onclick="setFilter('lic',this)">Licitações <span class="count-badge" id="cnt-lic"></span></button>
    <button class="btn btn-ma"     onclick="setFilter('aporte',this)">Aportes &amp; Investimentos <span class="count-badge" id="cnt-aporte"></span></button>
  </div>
</div>

<div class="main">

<div class="section-header" style="color:var(--blue-dark);border-color:var(--blue-mid);" id="sec-gestao">
  <h2>🏛️ Gestão Pública Municipal &amp; Radar Geral</h2>
  <span class="pill" style="background:var(--blue-light);color:var(--blue-dark);" id="cnt-gestao-sec"></span>
</div>
<div class="grid" id="grid-gestao"></div>

<div class="section-header" style="color:var(--teal-dark);border-color:var(--teal-mid);" id="sec-saude">
  <h2>🏥 Saúde Pública</h2>
  <span class="pill" style="background:var(--teal-light);color:var(--teal-dark);" id="cnt-saude-sec"></span>
</div>
<div class="grid" id="grid-saude"></div>

<div class="section-header" style="color:var(--blue-mid);border-color:var(--blue-mid);" id="sec-educ">
  <h2>📚 Educação Pública</h2>
  <span class="pill" style="background:var(--blue-light);color:var(--blue-mid);" id="cnt-educ-sec"></span>
</div>
<div class="grid" id="grid-educ"></div>

<div class="section-header" style="color:var(--purple-mid);border-color:var(--purple-mid);" id="sec-ia">
  <h2>🤖 IA para Governo</h2>
  <span class="pill" style="background:var(--purple-light);color:var(--purple-mid);" id="cnt-ia-sec"></span>
</div>
<div class="grid" id="grid-ia"></div>

<div class="section-header" style="color:var(--amber-dark);border-color:var(--amber-dark);" id="sec-proc">
  <h2>⚖️ Procuradorias (PGM/PGE/MP)</h2>
  <span class="pill" style="background:var(--amber-fill);color:var(--amber-dark);" id="cnt-proc-sec"></span>
</div>
<div class="grid" id="grid-proc"></div>

<div class="section-header" style="color:var(--lime-dark);border-color:var(--lime-dark);" id="sec-lic">
  <h2>📋 Licitações &amp; Compras Públicas</h2>
  <span class="pill" style="background:var(--lime-light);color:var(--lime-dark);" id="cnt-lic-sec"></span>
</div>
<div class="grid" id="grid-lic"></div>

<div class="section-header" style="color:var(--orange-dark);border-color:var(--orange-dark);" id="sec-aporte">
  <h2>💰 Aportes &amp; Investimentos</h2>
  <span class="pill" style="background:var(--orange-fill);color:var(--orange-dark);" id="cnt-aporte-sec"></span>
</div>
<div class="ma-table-wrap" id="grid-aporte">
  <table class="ma-table">
    <thead><tr>
      <th>Empresa / Investidor</th><th>Tipo</th><th>Valor</th>
      <th>Data</th><th>Segmento</th><th>Descrição</th>
    </tr></thead>
    <tbody id="aporte-tbody"></tbody>
  </table>
</div>

</div>

<div class="footer">
  Radar GovTech Brasil 2025/2026 · DLG · Uso interno · Gerado em {DATA_REF} · {VERSION}
</div>

<script>
const companies = {companies_json};
const aporteData = {aportes_json};

function getPorteClass(p) {{
  p = p.toLowerCase();
  if (p.includes("startup")) return "porte-startup";
  if (p.includes("scale"))  return "porte-scale";
  if (p.includes("pme"))    return "porte-pme";
  if (p.includes("médio") || p.includes("medio")) return "porte-medio";
  if (p.includes("grande") || p.includes("instituto") || p.includes("software")) return "porte-publico";
  return "porte-pme";
}}

function makeCard(c) {{
  const alerta = c.alerta || "";
  let tagClass = "";
  if (alerta.startsWith("Target") || alerta.startsWith("Ref") || alerta.startsWith("Vert") || alerta.startsWith("Parc") || alerta.startsWith("Sele") || alerta.startsWith("Oport")) tagClass = "ref";
  else if (alerta.startsWith("Watch") || alerta.startsWith("M&A")) tagClass = "watch";
  const tagHtml = alerta ? `<span class="card-${{tagClass}}-tag">${{alerta}}</span>` : "";
  const newBadge = c.isNew ? `<span class="card-new-tag">🆕 Novo</span>` : "";
  const nd = v => (!v || v === "N/D" || v === "N/D (Brasil)") ? "" : v;
  return `<div class="card ${{c.seg}}" data-seg="${{c.seg}}" data-search="${{(c.name+c.subseg+c.nota+c.loc+c.presenca).toLowerCase()}}">
    <div class="card-top">
      <div class="card-name">${{c.name}}</div>
      <div class="card-porte ${{getPorteClass(c.porte)}}">${{c.porte}}</div>
    </div>
    <div class="card-seg">${{c.subseg}}</div>
    <div class="card-meta">
      ${{nd(c.loc)     ? `<span class="meta-item"><span class="icon">📍</span>${{c.loc}}</span>` : ""}}
      ${{nd(c.fund)    ? `<span class="meta-item"><span class="icon">📅</span>${{c.fund}}</span>` : ""}}
      ${{nd(c.presenca)? `<span class="meta-item"><span class="icon">🌐</span>${{c.presenca}}</span>` : ""}}
      ${{nd(c.receita) ? `<span class="meta-item"><span class="icon">💰</span>${{c.receita}}</span>` : ""}}
    </div>
    <div class="card-nota">${{c.nota}}</div>
    <div style="margin-top:6px;">${{tagHtml}}${{newBadge}}</div>
  </div>`;
}}

function renderSection(gridId, cntSecId, seg) {{
  const arr = companies.filter(c => c.seg === seg);
  document.getElementById(gridId).innerHTML = arr.map(makeCard).join("");
  document.getElementById(cntSecId).textContent = arr.length + " empresas";
}}

function renderAportes() {{
  const tbody = document.getElementById("aporte-tbody");
  tbody.innerHTML = aporteData.map((m,i) => {{
    const tipoClass = m.tipo.includes("Exit") ? "tipo-exit" : "tipo-aporte";
    const bg = i%2===0 ? "" : ' style="background:#fafafa;"';
    return `<tr${{bg}}>
      <td style="font-weight:600;">${{m.tx}}</td>
      <td><span class="ma-tipo ${{tipoClass}}">${{m.tipo}}</span></td>
      <td style="font-weight:600;">${{m.valor}}</td>
      <td style="white-space:nowrap;">${{m.data}}</td>
      <td>${{m.seg}}</td>
      <td style="color:#444;">${{m.desc}}</td>
    </tr>`;
  }}).join("");
  const pill = document.getElementById("cnt-aporte-sec");
  if (pill) pill.textContent = aporteData.length + " registros";
  const badge = document.getElementById("cnt-aporte");
  if (badge) badge.textContent = aporteData.length;
}}

let currentFilter = "all";

function updateCounts() {{
  ["gestao","saude","educ","ia","proc","lic"].forEach(s => {{
    const el = document.getElementById("cnt-"+s);
    if (el) el.textContent = companies.filter(c=>c.seg===s).length;
  }});
  document.getElementById("cnt-all").textContent = companies.length;
}}

const sections = {{ gestao:["sec-gestao","grid-gestao"], saude:["sec-saude","grid-saude"], educ:["sec-educ","grid-educ"], ia:["sec-ia","grid-ia"], proc:["sec-proc","grid-proc"], lic:["sec-lic","grid-lic"], aporte:["sec-aporte","grid-aporte"] }};

function hideAllSections() {{ Object.values(sections).flat().forEach(id => {{ const el=document.getElementById(id); if(el) el.classList.add("hidden"); }}); }}
function showSection(id)    {{ const el=document.getElementById(id); if(el) el.classList.remove("hidden"); }}

function setFilter(f, btn) {{
  currentFilter = f;
  document.querySelectorAll(".filter-btns .btn").forEach(b => b.classList.remove("active"));
  btn.classList.add("active");
  hideAllSections();
  if (f === "all") {{ Object.values(sections).flat().forEach(showSection); }}
  else if (f === "aporte") {{ ["sec-aporte","grid-aporte"].forEach(showSection); }}
  else {{ ["sec-"+f,"grid-"+f].forEach(showSection); }}
  applySearch(document.getElementById("searchBox").value);
}}

function applySearch(q) {{
  const query = q.toLowerCase().trim();
  ["grid-gestao","grid-saude","grid-educ","grid-ia","grid-proc","grid-lic"].forEach(gid => {{
    const grid = document.getElementById(gid);
    if (!grid) return;
    grid.querySelectorAll(".card").forEach(card => {{
      const segMatch = currentFilter==="all" || currentFilter===card.dataset.seg;
      const qMatch = !query || (card.dataset.search||"").includes(query) || card.querySelector(".card-name").textContent.toLowerCase().includes(query);
      card.classList.toggle("hidden", !(segMatch && qMatch));
    }});
  }});
}}

document.getElementById("searchBox").addEventListener("input", function() {{ applySearch(this.value); }});

renderSection("grid-gestao","cnt-gestao-sec","gestao");
renderSection("grid-saude", "cnt-saude-sec", "saude");
renderSection("grid-educ",  "cnt-educ-sec",  "educ");
renderSection("grid-ia",    "cnt-ia-sec",    "ia");
renderSection("grid-proc",  "cnt-proc-sec",  "proc");
renderSection("grid-lic",   "cnt-lic-sec",   "lic");
renderAportes();
updateCounts();
</script>

<!-- LOGIN -->
<div id="login-overlay">
  <div id="login-box">
    <h2>Radar GovTech Brasil</h2>
    <p>DLG Tech — Acesso restrito</p>
    <input type="password" id="login-input" placeholder="Digite a senha" onkeydown="if(event.key==='Enter') checkLogin()" autofocus />
    <button onclick="checkLogin()">Entrar</button>
    <div id="login-error">Senha incorreta. Tente novamente.</div>
    <div id="login-logo">dlgtech.com.br</div>
  </div>
</div>
<script>
(function() {{
  if (sessionStorage.getItem('dlg_auth') === '1') {{ document.getElementById('login-overlay').style.display='none'; return; }}
  window.checkLogin = function() {{
    const pwd = document.getElementById('login-input').value;
    crypto.subtle.digest('SHA-256', new TextEncoder().encode(pwd)).then(b => {{
      const hash = Array.from(new Uint8Array(b)).map(x=>x.toString(16).padStart(2,'0')).join('');
      if (hash === 'a4ab044afdb741fec333da9498b48a42edf765c67fa259c8ed2916093d1d33a5') {{
        sessionStorage.setItem('dlg_auth','1');
        document.getElementById('login-overlay').style.display='none';
      }} else {{
        document.getElementById('login-error').style.display='block';
        document.getElementById('login-input').value='';
        document.getElementById('login-input').focus();
      }}
    }});
  }};
}})();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ HTML salvo: {output_path}")

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import os, sys
    sys.stdout.reconfigure(encoding="utf-8")
    base = os.path.dirname(os.path.abspath(__file__))

    print("Gerando Radar GovTech Brasil 2025/2026...\n")
    make_excel(os.path.join(base, "Radar_GovTechs_Brasil_2026.xlsx"))
    make_html(os.path.join(base, "index.html"))

    segs = {}
    for c in COMPANIES:
        segs[c["seg"]] = segs.get(c["seg"], 0) + 1
    unique = len({c["name"] for c in COMPANIES})

    print("\nResumo:")
    for seg, cfg in SEGMENTS.items():
        print(f"  {cfg['label']:<40} {segs.get(seg, 0)} empresas")
    print(f"\n  Total aparicoes : {sum(segs.values())}")
    print(f"  Empresas unicas : {unique}")
    print(f"  Aportes         : {len(APORTES)}")
    print("\nConcluido! Para publicar: python govtech_excel.py && firebase deploy --only hosting")
